import sys

import pandas as pd
from numpy import array, pad, max as n_max
from PyQt5 import QtCore
from PyQt5.QtWidgets import (QWidget, QTableWidget, QTableWidgetItem,
                             QApplication, QAbstractScrollArea, QSizePolicy,
                             QHeaderView, QLabel, QTextEdit, QTabWidget,
                             QGridLayout, QCalendarWidget, QPushButton,
                             QMessageBox,
                             QInputDialog, QFileDialog)
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import numbers
import re
from collections import defaultdict


class MappingExcelSheet(object):
    def __init__(self, sheet):
        self.excelformulas = ['SUM', 'COUNT', 'IF']
        self.min_cell = None
        self.excel_sheet = sheet
        self.set_map_excel()

    def set_map_excel(self):
        for i in range(1, self.excel_sheet.max_column):
            for j in range(1, self.excel_sheet.max_row):
                if self.excel_sheet.cell(i, j).value is None:
                    continue
                self.min_cell = [i, j]
                return

    def parse_formula(self, formula):
        formula = formula[1:] #lose the equal sign
        if formula.contains('('):
            formula_func = formula.split("(")[0] # Get sum or count or if
            if formula_func in self.excelformulas:
                within_func = formula.split("(")[1].split(')')[0]
                individual_cells = within_func.split(',')
        individual_cells = re.split('; |, |\+|\n|\*|-|/',formula)


class ExcelTable(object):
    UNNAMED = 'Unnamed'
    ADDTABLE = 'Add Team +'

    def __init__(self, widget, calendar, excel_file_path, tab_widget,
                 summary_sheet='Summary', sheet_name='Sheet1'):
        self.main_widget = widget
        self.calendar = calendar
        self.tab = tab_widget
        self.excel_file_path = excel_file_path
        self._activated_table = None
        self.summary_sheet = summary_sheet
        self.sheet_name = sheet_name
        self.index_names = []
        self.tables = []
        self.sheet_mapper = {}
        self.wb = load_workbook(self.excel_file_path)
        self.populate_excel_sheet(sheet_name=self.summary_sheet)
        self.populate_excel_sheet()
        self.add_table_to_tab(self.summary_sheet)
        self.set_summary_sheet_constant()
        self.add_table_to_tab(self.sheet_name)
        self.add_table_to_tab(self.ADDTABLE, True)
        self.tab.tabCloseRequested.connect(self.delete_tab)
        self.tab.currentChanged.connect(self.create_new_tab)
        self.tab.tabBarClicked.connect(self.tab_bar_clicked)
        self.map_all_sheet_name()

    def get_table_by_index(self, index):
        if 0 <= index < len(self.tables):
            return self.tables[index]
        return None

    def populate_excel_sheet(self, sheet_name=None):
        if sheet_name is None:
            sheet_name = self.sheet_name
        sheet_index = self.wb.sheetnames.index(sheet_name)
        sheet = self.wb.worksheets[sheet_index]
        mp_obj = MappingExcelSheet(sheet)
        self.sheet_mapper[sheet_name] = mp_obj

    def get_table_widget(self):
        table = QTableWidget()
        table.verticalHeader().setVisible(False)
        # table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        # table.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Maximum)
        # table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # table.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.horizontalHeader().setDefaultAlignment(
            QtCore.Qt.AlignHCenter | QtCore.Qt.Alignment(QtCore.Qt.TextWordWrap))
        table.verticalHeader().setDefaultAlignment(
            QtCore.Qt.AlignHCenter | QtCore.Qt.Alignment(QtCore.Qt.TextWordWrap))
        self.tables.append(table)
        return table

    def _groom_values(self, sheet):
        sheet_name = sheet.title
        if sheet.title == self.summary_sheet:
            return array([list(v) for v in sheet.values])
        values = []
        for row, v in enumerate(sheet.values):
            each_row = []
            for col, each_value in enumerate(v):
                if self.sheet_name in str(each_value):
                    each_value = str(each_value).replace(self.sheet_name, sheet_name)
                    sheet.cell(row + 1, col + 1).value = each_value
                each_row.append(each_value)
            values.append(each_row)
        return array(values)

    def populate_table(self, table, sheet_name, existing=False):
        if sheet_name not in self.sheet_mapper:
            if existing:
                sheet = self.wb[sheet_name]
            else:
                sheet = self.wb.copy_worksheet(self.wb[self.sheet_name])
            sheet.title = sheet_name
            mp_obj = MappingExcelSheet(sheet)
            self.sheet_mapper[sheet_name] = mp_obj
        cur_sheet_map = self.sheet_mapper[sheet_name]
        sheet = cur_sheet_map.excel_sheet
        values = self._groom_values(sheet)
        min_i, min_j = cur_sheet_map.min_cell
        table.setColumnCount(len(values[0]) - min_j + 1)
        table.setHorizontalHeaderLabels(values[min_i - 1, min_j -1:])
        table.setVerticalHeaderLabels(values[min_i - 1:, min_j - 1])
        table.setRowCount(len(values) - min_i + 1)
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        for i in range(cur_sheet_map.min_cell[0], len(values)):
            t = QTextEdit()
            t.setAlignment(QtCore.Qt.AlignCenter)
            # t.setWordWrap(True)
            ind_value = str('' if values[i, min_j - 1] is None
                            else values[i, min_j - 1])
            t.setText(ind_value)
            table.setIndexWidget(table.model().index(i - min_i, 0), t)
            for j in range(cur_sheet_map.min_cell[1], len(values[i])):
                text = str('' if values[i, j] is None else values[i, j])
                tw = QTableWidgetItem(text)
                table.setItem(i - min_i, j - min_j + 1, tw)
        table.cellChanged.connect(self.update_aggregation_df)
        table.cellClicked.connect(self.cell_activated)

    def map_all_sheet_name(self, sheet_names=None):
        self.map_names = defaultdict(list)
        if sheet_names is None:
            sheet_names = self.sheet_mapper.keys()
        for sheet_name in sheet_names:
            cur_sheet_map = self.sheet_mapper[sheet_name]
            values = array([list(v) for v in cur_sheet_map.excel_sheet.values])
            for i in range(cur_sheet_map.min_cell[0] - 1, len(values)):
                for j in range(cur_sheet_map.min_cell[1] - 1, len(values[i])):
                    value = '' if values[i, j] is None else str(values[i, j])
                    if sheet_name != self.summary_sheet:
                        if sheet_name in value:
                            self.map_names[sheet_name].append([sheet_name, i + 1, j + 1])
                        continue
                    for key in self.sheet_mapper.keys():
                        if key in value:
                            self.map_names[sheet_name].append([key, i + 1, j + 1])
        print(self.map_names)

    def add_table_to_tab(self, name, empty_table=False, existing=False):
        table = self.get_table_widget()
        if not empty_table:
            self.populate_table(table, name, existing=existing)
            # self.update_aggregation_df(None, [1, 2, 3, 4], -1)
        self.tab.addTab(table, name)

    def create_new_tab(self, table_index):
        if table_index != (len(self.tables) - 1):
            return
        table_name, ok = QInputDialog.getText(self.main_widget, 'Team Name',
                                              'Enter Team Name')
        if table_name == '' or ok is False:
            if table_index == (len(self.tables) - 1):
                self.tab.setCurrentIndex(0)
            return
        self._activated_table = None
        table = self.tables[table_index]
        self.populate_table(table, sheet_name=table_name)
        # self.update_aggregation_df(None, [1, 2, 3, 4], -1)
        self.tab.setTabText(table_index, table_name)
        self.add_new_row_in_summary(table_name)
        self.map_all_sheet_name()
        self.add_table_to_tab(self.ADDTABLE, True)

    def tab_bar_clicked(self, table_index):
        if (table_index == (len(self.tables) - 1) or
            table_index != self.tab.currentIndex() or
            self.tab.tabText(table_index) == 'Summary'):
            return
        table_name, ok = QInputDialog.getText(self.main_widget, 'Team Name',
                                              'Enter Team Name')
        if table_name == '' or ok is False:
            if table_index == (len(self.tables) - 1):
                self.tab.setCurrentIndex(0)
            return
        if table_index == self.tab.currentIndex():
            self.update_sheet_names(self.tab.tabText(table_index), table_name)
            self.tab.setTabText(table_index, table_name)
        self._activated_table = None

    def update_sheet_names(self, old_name, new_name):
        for sheet_key in self.map_names:
            cur_sheet_map = self.sheet_mapper[sheet_key]
            sheet = cur_sheet_map.excel_sheet
            m_i, m_j = cur_sheet_map.min_cell
            for key, i, j in self.map_names[sheet_key]:
                if key == old_name:
                    table = [t for i, t in enumerate(self.tables)
                              if self.tab.tabText(i) == sheet_key][0]
                    if i == m_i:
                        t_value = table.horizontalHeaderItem(j - m_j).text().replace(old_name, new_name)
                        table.horizontalHeaderItem(j-m_j).setText(t_value)
                    elif j == m_j:
                        t_value = table.indexWidget(table.model().index(
                            i - m_i - 1, 0)).toPlainText().replace(old_name, new_name)
                        table.indexWidget(table.model().index(
                            i - m_i - 1, 0)).setText(t_value)
                    else:
                        t_value = table.item(i - m_i - 1, j - m_j).text().replace(old_name, new_name)
                        table.item(i - m_i - 1, j - m_j).setText(t_value)
                    s_value = sheet.cell(i, j).value.replace(old_name, new_name)
                    sheet.cell(i, j).value = s_value
        if old_name in self.sheet_mapper:
          old_sheet_mapper_value = self.sheet_mapper.pop(old_name)
          old_sheet_mapper_value.excel_sheet.title = new_name
          self.sheet_mapper[new_name] = old_sheet_mapper_value
        if old_name == self.sheet_name:
            self.sheet_name = new_name
        sheet_map = self.sheet_mapper[new_name]
        sheet_map.title = new_name
        self.map_all_sheet_name()

    def set_summary_sheet_constant(self):
        summary_map = self.sheet_mapper.get('Summary')
        sheet = summary_map.excel_sheet
        self.sumamry_sheet_constant = []
        team_name = sheet.cell(sheet.max_row, 2).value
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(sheet.max_row, col).value
            if value is not None:
                value = value.replace(team_name, '--TEAM--')
            self.sumamry_sheet_constant.append(value)

    def add_new_row_in_summary(self, table_name):
        summary_map = self.sheet_mapper.get('Summary')
        table = self.tables[0]
        sheet = summary_map.excel_sheet
        m_i, m_j = summary_map.min_cell
        to_append = []
        table.setRowCount(table.rowCount() + 1)
        t = QTextEdit()
        t.setAlignment(QtCore.Qt.AlignCenter)
        # t.setWordWrap(True)
        for col, value in enumerate(self.sumamry_sheet_constant):
            if value is not None:
                value = value.replace('--TEAM--', table_name)
                if (col + 1 - m_j) == 0:
                    t.setText(value)
                    table.setIndexWidget(table.model().index(sheet.max_row - m_i, 0), t)
                else:
                    tw = QTableWidgetItem(value)
                    table.setItem(sheet.max_row - m_i, col + 1 - m_j, tw)
            to_append.append(value)
        sheet.append(to_append)

    def delete_row_in_summary(self, table_name):
        summary_map = self.sheet_mapper.get('Summary')
        table = self.tables[0]
        sheet = summary_map.excel_sheet
        m_i, m_j = summary_map.min_cell
        for row in range(1, sheet.max_row + 1):
          if sheet.cell(m_i + row, m_j).value == table_name:
            sheet.delete_rows(m_i + row)
            table.removeRow(row - 1)
            return

    def delete_tab(self, tab_index):
        sheet_name = self.tab.tabText(tab_index)
        if 'Add Team' in sheet_name:
          return
        self.tab.currentChanged.disconnect(self.create_new_tab)
        self.tab.tabBarClicked.disconnect(self.tab_bar_clicked)
        self.tab.setCurrentIndex(0)
        self.tables.pop(tab_index)
        self.tab.removeTab(tab_index)
        if sheet_name != self.sheet_name:
            self.wb.remove(self.wb[sheet_name])
        self.sheet_mapper.pop(sheet_name)
        self.map_all_sheet_name()
        self.delete_row_in_summary(sheet_name)
        self.tab.currentChanged.connect(self.create_new_tab)
        self.tab.tabBarClicked.connect(self.tab_bar_clicked)

    def cell_activated(self, row, col):
        self._activated_table = self.tab.currentIndex()

    def update_aggregation_df(self, row, col, table_index=None):
        print("Hit at ", row, col)
        if table_index is None:
            table_index = (self._activated_table
                           if self._activated_table is not None
                           else self.tab.currentIndex())
        table = self.tables[table_index]
        m = self.sheet_mapper[self.tab.tabText(table_index)]
        sheet = m.excel_sheet
        m_i, m_j = m.min_cell
        cell = sheet.cell(m_i + row + 1, m_j + col)
        txt, formt = self._get_dtyped_test(table.item(row, col).text())
        cell.value = txt
        if formt is not None:
          cell.number_format = formt
        self._activated_table = None

    @staticmethod
    def get_suffix(day):
        if 11 <= day <= 20:
            return 'th'
        if day % 10 == 1:
            return 'st'
        if day % 10 == 2:
            return 'nd'
        if day % 10 == 3:
            return 'rd'
        return 'th'

    def show_date(self, date_selected):
        stagnant_row = 2
        col_maps = {}
        for i, (sheet_name, cur_sheet_mapper) in enumerate(self.sheet_mapper.items()):
            if sheet_name == 'Summary':
                continue
            excel_sheet = cur_sheet_mapper.excel_sheet
            min_i, min_j = cur_sheet_mapper.min_cell
            for col in range(1, excel_sheet.max_column + 1):
                value = excel_sheet.cell(stagnant_row, col).value
                if not value:
                    continue
                if col not in col_maps:
                    date_end = date_selected.addDays(6)
                    start_day = date_selected.day()
                    end_day = date_end.day()
                    week_str = (date_selected.shortMonthName(date_selected.month())
                                + " " + str(start_day) + self.get_suffix(start_day)
                                + " - " + date_end.shortMonthName(date_end.month())
                                + " " + str(end_day) + self.get_suffix(end_day))
                    col_maps[col] = week_str
                week_str = col_maps.get(col)
                excel_sheet.cell(stagnant_row, col).value = week_str
                for table in self.tables[1:-1]:
                    week_item = table.item(stagnant_row - 1 - min_i, col-min_j)
                    week_item.setText(week_str)
                if col not in col_maps:
                    date_selected = date_selected.addDays(7)
        self.calendar.setVisible(False)


    def _get_dtyped_test(self, value):
        try:
            if float(value) == int(value):
                return int(value), numbers.FORMAT_NUMBER
            else:
                return float(value), numbers.FORMAT_NUMBER
        except Exception as ex:
            return value, None
        return value, None

    def clear_all(self):
        self.tab.clear()
        self.tables = []
        self.sheet_mapper = {}
        self._activated_table = None
        self.excel_file_path = r''
        self.sheet_name = ''

    def import_from_excel(self, file_path):
        response = QMessageBox.question(self.main_widget, 'Data Clear',
                                        "Are you sure to clear the existing data")
        self.wb = load_workbook(file_path)
        if response == QMessageBox.Yes:
            self.clear_all()
            self.tab.currentChanged.disconnect(self.create_new_tab)
            self.tab.tabBarClicked.disconnect(self.tab_bar_clicked)
            self.excel_file_path = file_path
            self.sheet_name = self.wb.sheetnames[1]
            self.populate_excel_sheet(sheet_name=self.summary_sheet)
            self.add_table_to_tab(self.summary_sheet, existing=True)
            for sheet in self.wb.sheetnames[1:]:
                self.populate_excel_sheet()
                self.add_table_to_tab(sheet, existing=True)
            self.set_summary_sheet_constant()
            self.map_all_sheet_name()
            self.add_table_to_tab(self.ADDTABLE, True)
            self.tab.currentChanged.connect(self.create_new_tab)
            self.tab.tabBarClicked.connect(self.tab_bar_clicked)

    def export_to_excel(self, file_path):
        if file_path == '':
            QMessageBox.warning(self.main_widget, "failure", "Not a valid path")
            return
        if self.sheet_name not in self.sheet_mapper:
            self.wb.remove(self.wb[self.sheet_name])
        self.wb.save(file_path)
        QMessageBox.information(self.main_widget, "Success", "File Exported Successfully!!")


def main_widget():
    win = QWidget()
    tab = QTabWidget()
    tab.setTabsClosable(True)

    cal = QCalendarWidget()
    cal.setGridVisible(True)
    cal.move(20, 20)
    cal.setVisible(False)
    excel_path = r'F:\python excel code\support_roaster_format.xlsx'
    et = ExcelTable(win, cal, excel_path, tab, summary_sheet='Summary',
                    sheet_name='Prudents')
    cal.clicked.connect(et.show_date)

    def open_calendar(args):
        cal.setVisible(True)

    def file_save_as(args):
        filePath, _ = QFileDialog.getSaveFileName(win, "Export to Excel", "",
            "Excel Workbook (*.xlsx);;Excel 97-2003 Workbook (*.xls);;CSV UTF-8 (Comma delimited) (*.csv)")
        et.export_to_excel(filePath)

    def file_open(args):
        filePath, _ = QFileDialog.getOpenFileName(win, "Import from Excel", "",
            "Excel Workbook (*.xlsx);;Excel 97-2003 Workbook (*.xls);;CSV UTF-8 (Comma delimited) (*.csv)")
        et.import_from_excel(filePath)

    set_data_button = QPushButton("Set First Week")
    set_data_button.setFixedSize(200, 50)
    set_data_button.clicked.connect(open_calendar)

    load_existing = QPushButton("Load Existing")
    load_existing.setFixedSize(200, 50)
    load_existing.clicked.connect(file_open)

    save_data_button = QPushButton("Export to Excel")
    save_data_button.setFixedSize(200, 50)
    save_data_button.clicked.connect(file_save_as)

    layout = QGridLayout()
    layout.addWidget(set_data_button, 0, 0)
    layout.addWidget(cal, 0, 1)
    layout.addWidget(load_existing, 0, 2)
    layout.addWidget(tab, 1, 0, 1, 0)
    layout.addWidget(save_data_button, 2, 0)
    win.setLayout(layout)
    return win, cal, et


if __name__ == '__main__':
    app = QApplication(sys.argv)
    win, cal, et = main_widget()
    cal.clicked.connect(et.show_date)
    # win.show()
    win.showMaximized()
    app.exec_()

